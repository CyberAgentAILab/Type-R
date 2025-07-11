import argparse
import os

import numpy as np
import openpyxl
import PIL.Image
import seaborn as sns
import yaml
from hydra.utils import call
from loguru import logger
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, pixels_to_points
from type_r.util.structure import WordMapping

from type_r_app.launcher.typo_correction import load_word_mapping
from type_r_app.visualizer import Visualizer

palettes = sns.color_palette(None, 15)


def save_vis_polygon(img: PIL.Image.Image, word_mapping: WordMapping, save_name: str):
    vis = Visualizer(img)
    for i, polygon in enumerate(word_mapping.polygons):
        vis.draw_polygon(segment=np.array(polygon), color=palettes[i % len(palettes)])
    vis.output.save(save_name)


def wb_config(wb, sheet_name, EXEL_WIDTH_UNIT):
    wb[sheet_name].cell(1, 1).value = "prompt"
    wb[sheet_name].cell(1, 2).value = "raw"
    wb[sheet_name].cell(1, 3).value = "layout correction"
    wb[sheet_name].cell(1, 4).value = "typo correction"

    wb[sheet_name].column_dimensions["A"].width = EXEL_WIDTH_UNIT
    wb[sheet_name].column_dimensions["B"].width = EXEL_WIDTH_UNIT
    wb[sheet_name].column_dimensions["C"].width = EXEL_WIDTH_UNIT
    wb[sheet_name].column_dimensions["D"].width = EXEL_WIDTH_UNIT
    return wb


def add_image(ws, row, col, fn):
    image = openpyxl.drawing.image.Image(fn)
    margin = 5
    img_size_w = 300
    img_size_h = 300
    col_offset = pixels_to_EMU(margin)
    row_offset = pixels_to_EMU(margin)
    size_ext = XDRPositiveSize2D(pixels_to_EMU(img_size_w), pixels_to_EMU(img_size_h))
    maker = AnchorMarker(col=col, colOff=col_offset, row=row, rowOff=row_offset)
    image.anchor = OneCellAnchor(_from=maker, ext=size_ext)
    ws.add_image(image)


def main(args):
    result_dir = args.result_dir
    with open(args.config_path, "r", encoding="utf-8") as file:
        config = yaml.safe_load(file)
    hfds = call(config)()()

    name = f"{result_dir}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    sheet_name = "res"
    ws.title = sheet_name
    img_size_w = 300
    img_size_h = 300
    EXEL_WIDTH_UNIT = (img_size_w + 5 * 2) / 8
    EXEL_HEIGHT_UNIT = pixels_to_points(img_size_h + 5 * 2)

    wb = wb_config(wb, sheet_name, EXEL_WIDTH_UNIT)

    ws.title = sheet_name

    os.makedirs(os.path.join(result_dir, "layout_corrected_img_vis"), exist_ok=True)

    for i, element in enumerate(hfds):
        dataset = element["dataset_name"]
        idx = element["id"]
        prompt = element["prompt"]
        wb[sheet_name].row_dimensions[i + 2].height = EXEL_HEIGHT_UNIT

        # prompt
        wb[sheet_name].cell(i + 2, 1).value = prompt
        wb[sheet_name].cell(i + 2, 1).alignment = openpyxl.styles.Alignment(
            wrapText=True
        )

        # t2i result
        add_image(
            ws,
            i + 1,
            1,
            os.path.join(result_dir, "ref_img", f"{dataset}_{str(idx)}.jpg"),
        )
        # layout correction result
        word_mapping = load_word_mapping(f"{result_dir}/word_mapping", dataset, idx)
        img = PIL.Image.open(
            os.path.join(
                result_dir, "layout_corrected_img", f"{dataset}_{str(idx)}.jpg"
            )
        )
        save_name = os.path.join(
            result_dir, "layout_corrected_img_vis", f"{dataset}_{str(idx)}.jpg"
        )
        save_vis_polygon(img, word_mapping, save_name)
        add_image(
            ws,
            i + 1,
            2,
            save_name,
        )

        # typo correction result
        add_image(
            ws,
            i + 1,
            3,
            os.path.join(result_dir, "typo_corrected_img", f"{dataset}_{str(idx)}.jpg"),
        )
        logger.info(f"{i}/{len(hfds)}")
    logger.info(f"saved to {name}")
    wb.save(name)


def get_args():
    parser = argparse.ArgumentParser(description="This is sample argparse script")
    parser.add_argument(
        "--result_dir",
        default="results/marioevalbench_best",
        type=str,
        help="The path of by-products generated by the Type-R process.",
    )
    parser.add_argument(
        "--config_path",
        default="src/type_r_app/config/dataset/marioeval.yaml",
        type=str,
        help="Path to the YAML file for the Hugging Face dataset configuration.",
    )

    return parser.parse_args()


if __name__ == "__main__":
    args = get_args()
    main(args)
